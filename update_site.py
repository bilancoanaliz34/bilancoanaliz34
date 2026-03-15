# -*- coding: utf-8 -*-
"""
update_site.py
1-Borsa Calisma.xlsx -> veri sayfasini okur
-> index.html icindeki VDATA bolumunu gunceller
-> git add / commit / push yapar
"""

import sys
import re
import gzip
import base64
import json
import subprocess
from pathlib import Path
from datetime import datetime

# stdout'u UTF-8 yap (Windows'ta cp1254 sorunu)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Sabitler ────────────────────────────────────────────────────
REPO_DIR   = Path(__file__).parent
EXCEL_FILE = REPO_DIR / "1-Borsa \u00c7al\u0131\u015fma.xlsx"   # 1-Borsa Çalışma.xlsx
HTML_FILE  = REPO_DIR / "index.html"
SHEET_NAME = "veri"

# JS parseWorkbook ile birebir eslesen sutun adlari
COL_MAP = {
    "hisse"   : "Hisse",
    "donem"   : "D\u00f6nem",                        # Dönem
    "sektor"  : "Sekt\u00f6r",                       # Sektör
    "firma"   : "F\u0130RMA \u0130SM\u0130",          # FİRMA İSMİ
    "bilTarih": "Bilan\u00e7o Tarihi",               # Bilanço Tarihi
    "fiyat"   : "Hisse Fiyat\u0131",                 # Hisse Fiyatı
    "adet"    : "Hisse Say\u0131s\u0131",            # Hisse Sayısı
    "pd"      : "Piyasa De\u011feri",                # Piyasa Değeri
    "fd"      : "Firma De\u011feri",                 # Firma Değeri
    "donenV"  : "D\u00f6nen Varl\u0131klar",         # Dönen Varlıklar
    "duranV"  : "Duran Varl\u0131klar",              # Duran Varlıklar
    "kvYuk"   : "K\u0131sa Vadeli Y\u00fck\u00fcml\u00fcl\u00fckler",  # Kısa Vadeli Yükümlülükler
    "uvYuk"   : "Uzun Vadeli Y\u00fck\u00fcml\u00fcl\u00fckler",       # Uzun Vadeli Yükümlülükler
    "oz"      : "\u00d6zkaynak",                     # Özkaynak
    "brutKar" : "Br\u00fct Kar",                     # Brüt Kar
    "faalKar" : "Faliyet Kar\u0131 Zarar\u0131",     # Faliyet Karı Zararı
    "favok"   : "FAV\u00d6K",                        # FAVÖK
    "netBorc" : "NET BOR\u00c7",                     # NET BORÇ
    "netKar"  : "NET KAR YILLIK",
    "satislar": "NET SATI\u015eLAR",                 # NET SATIŞLAR
    "roe"     : "\u00d6zkaynak Karl\u0131l\u0131\u011f\u0131",  # Özkaynak Karlılığı
    "ol"      : "OLUMLU",
    "no"      : "N\u00d6TR",                         # NÖTR
    "ols"     : "OLUMSUZ",
    "yo"      : "YO",
    "pddd"    : "PD/DD",
    "fk"      : "F/K",
    "fdFavok" : "FD/FAV\u00d6K",                     # FD/FAVÖK
    "pdNfk"   : "PD/NFK",
    "fdNs"    : "FD/NS",
    "nfkPd"   : "NFK/PD",
    "nbFavok" : "NETBOR\u00c7/FAV\u00d6K",           # NETBORÇ/FAVÖK
    "efk"     : "EFK",
    "hbk"     : "HBK",
}

# ── Yardimci fonksiyonlar ────────────────────────────────────────
def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return None if f != f else f  # NaN
    except (TypeError, ValueError):
        return None

def numi(v):
    n = num(v)
    return None if n is None else int(round(n))

def sort_key(p):
    parts = str(p).split("/")
    y = int(parts[0]) if parts[0].isdigit() else 0
    m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
    return y * 100 + m

def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return None

# ── Excel okuma ──────────────────────────────────────────────────
def parse_excel(path: Path) -> dict:
    print("Excel okunuyor:", path.name)
    wb = openpyxl.load_workbook(str(path), data_only=True)

    sheet_lower = {s.lower(): s for s in wb.sheetnames}
    if SHEET_NAME not in sheet_lower:
        raise ValueError("'veri' sayfasi bulunamadi. Mevcut: " + str(wb.sheetnames))
    ws = wb[sheet_lower[SHEET_NAME]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError("veri sayfasinda yeterli satir yok")

    # Baslik satiri (2. satir = index 1)
    hdr = rows[1]
    ci = {}
    for idx, cell in enumerate(hdr):
        if cell is not None:
            ci[str(cell).strip()] = idx

    def get_idx(key):
        col_name = COL_MAP.get(key)
        if col_name and col_name in ci:
            return ci[col_name]
        return -1

    result = {}
    skipped = 0

    for row in rows[2:]:
        if not row:
            continue
        hisse_idx = get_idx("hisse")
        donem_idx = get_idx("donem")
        if hisse_idx < 0 or donem_idx < 0:
            continue

        ticker = str(row[hisse_idx] or "").strip().upper()
        donem  = str(row[donem_idx]  or "").strip()

        if not ticker or not donem:
            skipped += 1
            continue
        if not re.match(r"^\d{4}/\d+$", donem):
            skipped += 1
            continue

        if ticker not in result:
            firma_idx = get_idx("firma")
            sektor_idx = get_idx("sektor")
            result[ticker] = {
                "ticker"      : ticker,
                "company"     : str(row[firma_idx] if firma_idx >= 0 else None or ticker),
                "sector"      : str(row[sektor_idx] if sektor_idx >= 0 else None or ""),
                "puan"        : None,
                "bilancoTarih": None,
                "periods"     : [],
                "rows"        : {},
            }

        r = result[ticker]
        if donem not in r["periods"]:
            r["periods"].append(donem)

        # Bilanc tarihi — en guncel tarih
        bil_idx = get_idx("bilTarih")
        bt = fmt_date(row[bil_idx] if bil_idx >= 0 and bil_idx < len(row) else None)
        if bt and (not r["bilancoTarih"] or bt > r["bilancoTarih"]):
            r["bilancoTarih"] = bt

        def g(key):
            idx = get_idx(key)
            return row[idx] if 0 <= idx < len(row) else None

        donen_v = num(g("donenV"))
        duran_v = num(g("duranV"))
        tv = (donen_v + duran_v) if (donen_v is not None and duran_v is not None) else None

        yo_float = num(g("yo"))   # Excel'de 0-1 aralik (ornek: 0.54)
        yo_stored = round(yo_float * 100) if yo_float is not None else None  # 54

        r["rows"][donem] = {
            "fd"      : num(g("fd")),
            "pd"      : num(g("pd")),
            "fiyat"   : num(g("fiyat")),
            "adet"    : num(g("adet")),
            "donenV"  : donen_v,
            "duranV"  : duran_v,
            "kvYuk"   : num(g("kvYuk")),
            "uvYuk"   : num(g("uvYuk")),
            "tv"      : tv,
            "oz"      : num(g("oz")),
            "brutKar" : num(g("brutKar")),
            "faalKar" : num(g("faalKar")),
            "favok"   : num(g("favok")),
            "netBorc" : num(g("netBorc")),
            "netKar"  : num(g("netKar")),
            "satislar": num(g("satislar")),
            "roe"     : num(g("roe")),
            "ol"      : numi(g("ol")),
            "no"      : numi(g("no")),
            "ols"     : numi(g("ols")),
            "yo"      : yo_stored,
            "pddd"    : num(g("pddd")),
            "fk"      : num(g("fk")),
            "fdFavok" : num(g("fdFavok")),
            "pdNfk"   : num(g("pdNfk")),
            "fdNs"    : num(g("fdNs")),
            "nfkPd"   : num(g("nfkPd")),
            "nbFavok" : num(g("nbFavok")),
            "efk"     : num(g("efk")),
            "hbk"     : num(g("hbk")),
        }

    # Donemleri sirala, puan = son donemin yo degeri
    for ticker_key, r in result.items():
        r["periods"].sort(key=sort_key, reverse=True)
        if r["periods"]:
            p0 = r["periods"][0]
            yo_p0 = r["rows"].get(p0, {}).get("yo")
            if yo_p0 is not None:
                r["puan"] = yo_p0

    print("  ->", len(result), "hisse,", skipped, "satir atlandi")
    return result

# ── Sikistir + base64 ───────────────────────────────────────────
def to_b64_gzip(data: dict) -> str:
    json_bytes = json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    compressed = gzip.compress(json_bytes, compresslevel=9)
    return base64.b64encode(compressed).decode("ascii")

# ── index.html guncelle ──────────────────────────────────────────
def update_html(b64_data: str, html_path: Path):
    print("HTML guncelleniyor:", html_path.name)
    html = html_path.read_text(encoding="utf-8")

    pattern = r'(<script\s+id="VDATA"[^>]*>)(.*?)(</script>)'
    match = re.search(pattern, html, re.DOTALL)
    if not match:
        raise ValueError('index.html icinde <script id="VDATA"> bulunamadi')

    new_html = html[:match.start()] + \
               match.group(1) + b64_data + match.group(3) + \
               html[match.end():]

    html_path.write_text(new_html, encoding="utf-8")
    print("  -> VDATA guncellendi (" + f"{len(b64_data):,}" + " karakter)")

# ── Git islemleri ────────────────────────────────────────────────
def run_git(*args):
    result = subprocess.run(
        ["git"] + list(args),
        cwd=str(REPO_DIR),
        capture_output=True, text=True, encoding="utf-8", errors="replace"
    )
    if result.stdout.strip():
        print("  git:", result.stdout.strip())
    if result.stderr.strip():
        print("  git stderr:", result.stderr.strip())
    if result.returncode != 0:
        raise RuntimeError("git " + " ".join(args) + " basarisiz (kod " + str(result.returncode) + ")")
    return result.stdout.strip()

def git_push():
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    print("Git: dosyalar ekleniyor...")
    run_git("add", "index.html")

    status = subprocess.run(
        ["git", "status", "--porcelain", "index.html"],
        cwd=str(REPO_DIR), capture_output=True, text=True, encoding="utf-8"
    )
    if not status.stdout.strip():
        print("  -> Degisiklik yok, commit atlandi.")
        return

    print("Git: commit yapiliyor...")
    run_git("commit", "-m", "veri guncelle " + now)

    print("Git: push yapiliyor...")
    run_git("push", "origin", "HEAD:main")
    print("  -> Push tamamlandi.")

# ── Ana akis ────────────────────────────────────────────────────
def main():
    if not EXCEL_FILE.exists():
        print("HATA: Excel dosyasi bulunamadi ->", str(EXCEL_FILE))
        sys.exit(1)
    if not HTML_FILE.exists():
        print("HATA: index.html bulunamadi ->", str(HTML_FILE))
        sys.exit(1)

    data = parse_excel(EXCEL_FILE)
    b64  = to_b64_gzip(data)
    update_html(b64, HTML_FILE)
    git_push()
    print("Tamamlandi.")

if __name__ == "__main__":
    main()
